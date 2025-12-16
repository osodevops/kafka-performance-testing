#!/usr/bin/env python3
"""
Kafka Performance Excel Report Generator (PRD v2 Implementation)

Generates comprehensive, graph-driven Excel reports from parsed Kafka performance
test results. Creates a 10-sheet workbook with 30+ visualizations including:
- Dashboard with KPI summary and key charts
- Throughput/Latency deep-dive analysis
- THE CRITICAL "knee" chart showing saturation point
- Configuration heatmaps for multi-parameter optimization
- Scaling performance degradation curves
- Scored recommendations by use case

Usage:
    python generate_excel_report.py <parsed_json> <output_xlsx>
    python generate_excel_report.py ./results/parsed_data/parsed_results.json ./results/reports/kafka_perf_report.xlsx

Sheets Generated:
    1. Dashboard - Executive summary with 4 key charts
    2. Throughput Analysis - Deep dive into throughput optimization
    3. Latency Analysis - Percentile distributions and trends
    4. Trade-off Analysis - THE KNEE CHART + multi-dimensional analysis
    5. Scaling Performance - Producer/consumer scaling degradation
    6. Configuration Heatmap - Multi-dimensional parameter analysis
    7. Message Size Impact - Size efficiency curves
    8. Acks Comparison - Durability vs performance trade-offs
    9. Raw Data - Complete data for pivot tables
    10. Recommendations - Scored configurations by use case
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from statistics import mean, stdev

try:
    import openpyxl
    from openpyxl.chart import (
        LineChart, BarChart, ScatterChart, Reference, AreaChart,
        RadarChart, BubbleChart, Series
    )
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.trendline import Trendline
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
except ImportError:
    print("[ERROR] openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("[ERROR] pandas is required. Install with: pip install pandas")
    sys.exit(1)

try:
    import numpy as np
except ImportError:
    np = None  # Optional, will use pure Python fallbacks


class KafkaPerformanceReporter:
    """
    Excel report generator for Kafka performance test results.

    Creates comprehensive workbooks with 30+ visualizations including:
    - Dashboard with executive KPIs and key charts
    - THE CRITICAL "knee" chart (latency vs throughput saturation curve)
    - Configuration heatmaps for multi-parameter optimization
    - Scaling degradation analysis
    - Scored recommendations by use case
    """

    # ==========================================================================
    # STYLE DEFINITIONS (PRD Color Scheme)
    # ==========================================================================

    # Header styles
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(bold=True, size=18, color="1F4E78")
    SUBTITLE_FONT = Font(bold=True, size=14, color="2E75B6")
    SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
    METRIC_FONT = Font(bold=True, size=11)

    # Status fills
    GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # PRD-specified color scheme for acks settings
    ACKS_COLORS = {
        '0': 'FF6B6B',      # Red - fast but risky
        '1': 'FFD93D',      # Yellow - balanced
        'all': '6BCB77',    # Green - safe but slow
        '-1': '6BCB77',     # Same as 'all'
    }

    # Throughput zone colors
    ZONE_COLORS = {
        'excellent': '2F5740',  # Dark green
        'good': 'A4DE6C',       # Light green
        'ok': 'FFE66D',         # Yellow
        'poor': 'FF6B6B',       # Orange-red
        'critical': '8B0000',   # Dark red
    }

    # Chart style settings
    CHART_STYLE = 13  # Modern Excel chart style
    CHART_BG_COLOR = 'F5F5F5'

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    THICK_BORDER = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    def __init__(self, parsed_json_file: str, output_xlsx: str, verbose: bool = False):
        """
        Initialize the reporter.

        Args:
            parsed_json_file: Path to parsed JSON results file
            output_xlsx: Path for output Excel file
            verbose: Enable verbose output
        """
        self.output_file = Path(output_xlsx)
        self.verbose = verbose

        # Load parsed data
        self.data = self._load_parsed_data(parsed_json_file)

        # Create workbook
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Organize data by test type
        self.producer_results = [r for r in self.data if r.get('test_type') == 'producer']
        self.consumer_results = [r for r in self.data if r.get('test_type') == 'consumer']

        self.log(f"Loaded {len(self.producer_results)} producer results")
        self.log(f"Loaded {len(self.consumer_results)} consumer results")

        # Create DataFrames for analysis
        self.producer_df = self._create_producer_dataframe()
        self.consumer_df = self._create_consumer_dataframe()

        # Pre-calculate scores and statistics
        self.scores = self._calculate_all_scores()
        self.stats = self._calculate_statistics()

    def _create_producer_dataframe(self) -> pd.DataFrame:
        """Convert producer results to DataFrame for analysis and charting."""
        if not self.producer_results:
            return pd.DataFrame()

        rows = []
        for r in self.producer_results:
            config = r.get('configuration', {})
            metrics = r.get('metrics', {})
            rows.append({
                'scenario': r.get('scenario', 'unknown'),
                'acks': str(config.get('acks', '')),
                'batch_size': config.get('batch_size', 0),
                'linger_ms': config.get('linger_ms', 0),
                'compression_type': config.get('compression_type', config.get('compression', 'none')),
                'record_size': config.get('record_size', 1024),
                'num_producers': config.get('num_producers', 1),
                'throughput_rps': metrics.get('throughput_rps', 0),
                'throughput_mb': metrics.get('throughput_mb', 0),
                'avg_latency_ms': metrics.get('avg_latency_ms', 0),
                'max_latency_ms': metrics.get('max_latency_ms', 0),
                'p50_ms': metrics.get('p50_ms', 0),
                'p95_ms': metrics.get('p95_ms', 0),
                'p99_ms': metrics.get('p99_ms', 0),
                'p999_ms': metrics.get('p999_ms', 0),
            })
        return pd.DataFrame(rows)

    def _create_consumer_dataframe(self) -> pd.DataFrame:
        """Convert consumer results to DataFrame for analysis and charting."""
        if not self.consumer_results:
            return pd.DataFrame()

        rows = []
        for r in self.consumer_results:
            config = r.get('configuration', {})
            metrics = r.get('metrics', {})
            rows.append({
                'scenario': r.get('scenario', 'unknown'),
                'fetch_min_bytes': config.get('fetch_min_bytes', 0),
                'max_poll_records': config.get('max_poll_records', 500),
                'num_consumers': config.get('num_consumers', 1),
                'throughput_mb_sec': metrics.get('throughput_mb_sec', 0),
                'throughput_msg_sec': metrics.get('throughput_msg_sec', 0),
                'data_consumed_mb': metrics.get('data_consumed_mb', 0),
                'rebalance_time_ms': metrics.get('rebalance_time_ms', 0),
                'fetch_time_ms': metrics.get('fetch_time_ms', 0),
                'fetch_mb_sec': metrics.get('fetch_mb_sec', 0),
            })
        return pd.DataFrame(rows)

    def _calculate_all_scores(self) -> Dict[str, Any]:
        """
        Calculate performance scores for all configurations.

        Scoring methodology (from PRD):
        - Throughput Score = (test_throughput / max_observed) * 100
        - Latency Score = 100 - ((test_latency / max_observed) * 100)
        - Consistency Score = (1 - stddev/mean) * 100 (coefficient of variation)

        Returns dict with scored results for each use case weighting.
        """
        if self.producer_df.empty:
            return {}

        df = self.producer_df.copy()

        # Calculate normalization values
        max_throughput = df['throughput_mb'].max() if df['throughput_mb'].max() > 0 else 1
        max_latency = df['avg_latency_ms'].max() if df['avg_latency_ms'].max() > 0 else 1

        # Calculate individual scores
        df['throughput_score'] = (df['throughput_mb'] / max_throughput) * 100
        df['latency_score'] = 100 - ((df['avg_latency_ms'] / max_latency) * 100)

        # Consistency score based on p99/avg ratio (lower is more consistent)
        df['consistency_score'] = df.apply(
            lambda row: max(0, 100 - ((row['p99_ms'] / row['avg_latency_ms'] - 1) * 50))
            if row['avg_latency_ms'] > 0 else 50,
            axis=1
        )

        # Calculate weighted scores for different use cases
        # Max Throughput: 40% throughput + 20% latency + 40% consistency
        df['score_max_throughput'] = (
            0.4 * df['throughput_score'] +
            0.2 * df['latency_score'] +
            0.4 * df['consistency_score']
        )

        # Balanced: 30% throughput + 35% latency + 35% consistency
        df['score_balanced'] = (
            0.3 * df['throughput_score'] +
            0.35 * df['latency_score'] +
            0.35 * df['consistency_score']
        )

        # Durability: 20% throughput + 35% latency + 45% consistency
        df['score_durability'] = (
            0.2 * df['throughput_score'] +
            0.35 * df['latency_score'] +
            0.45 * df['consistency_score']
        )

        return {
            'scored_df': df,
            'max_throughput': max_throughput,
            'max_latency': max_latency,
            'best_throughput_idx': df['throughput_mb'].idxmax() if not df.empty else None,
            'best_latency_idx': df['avg_latency_ms'].idxmin() if not df.empty else None,
            'best_balanced_idx': df['score_balanced'].idxmax() if not df.empty else None,
        }

    def _calculate_statistics(self) -> Dict[str, Any]:
        """Calculate aggregate statistics for dashboard and analysis."""
        stats = {
            'producer': {},
            'consumer': {},
            'scaling': {},
        }

        if not self.producer_df.empty:
            df = self.producer_df
            stats['producer'] = {
                'max_throughput_mb': df['throughput_mb'].max(),
                'max_throughput_rps': df['throughput_rps'].max(),
                'min_latency_ms': df['avg_latency_ms'].min(),
                'avg_throughput_mb': df['throughput_mb'].mean(),
                'avg_latency_ms': df['avg_latency_ms'].mean(),
                'test_count': len(df),
            }

            # Scaling analysis - group by num_producers
            if 'num_producers' in df.columns:
                scaling_data = df.groupby('num_producers').agg({
                    'throughput_mb': 'mean',
                    'avg_latency_ms': 'mean',
                }).reset_index()
                stats['scaling']['producer'] = scaling_data.to_dict('records')

        if not self.consumer_df.empty:
            df = self.consumer_df
            stats['consumer'] = {
                'max_throughput_mb': df['throughput_mb_sec'].max(),
                'max_throughput_rps': df['throughput_msg_sec'].max(),
                'avg_throughput_mb': df['throughput_mb_sec'].mean(),
                'test_count': len(df),
            }

        return stats

    def log(self, message: str) -> None:
        """Print message if verbose mode is enabled."""
        if self.verbose:
            print(f"[INFO] {message}")

    def _load_parsed_data(self, json_file: str) -> List[Dict[str, Any]]:
        """Load parsed results from JSON file."""
        json_path = Path(json_file)

        if not json_path.exists():
            print(f"[ERROR] JSON file not found: {json_file}")
            sys.exit(1)

        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Handle both single result and list of results
        if isinstance(data, dict):
            return [data]
        return data

    def generate_report(self) -> str:
        """
        Generate the complete Excel report with 10 sheets and 30+ visualizations.

        Sheets created (per PRD v2):
        1. Dashboard - Executive summary with key charts
        2. Throughput Analysis - Deep dive into throughput optimization
        3. Latency Analysis - Percentile distributions and trends
        4. Trade-off Analysis - THE KNEE CHART + multi-dimensional analysis
        5. Scaling Performance - Producer/consumer scaling degradation
        6. Configuration Heatmap - Multi-dimensional parameter analysis
        7. Message Size Impact - Size efficiency curves
        8. Acks Comparison - Durability vs performance trade-offs
        9. Raw Data - Complete data for pivot tables
        10. Recommendations - Scored configurations by use case

        Returns:
            Path to the generated Excel file
        """
        self.log("Generating comprehensive Excel report (PRD v2)...")
        self.log(f"Creating 10 sheets with 30+ visualizations...")

        # Sheet 1: Dashboard (Executive Summary)
        self.create_dashboard_sheet()

        # Sheet 2: Throughput Analysis
        self.create_throughput_analysis_sheet()

        # Sheet 3: Latency Analysis
        self.create_latency_analysis_sheet()

        # Sheet 4: Trade-off Analysis (THE CRITICAL SHEET)
        self.create_tradeoff_analysis_sheet()

        # Sheet 5: Scaling Performance
        self.create_scaling_performance_sheet()

        # Sheet 6: Configuration Heatmap
        self.create_heatmap_sheet()

        # Sheet 7: Message Size Impact
        self.create_message_size_sheet()

        # Sheet 8: Acks Comparison
        self.create_acks_comparison_sheet()

        # Sheet 9: Raw Data
        self.create_raw_data_sheet()

        # Sheet 10: Recommendations
        self.create_recommendations_sheet()

        # Ensure output directory exists
        self.output_file.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        self.wb.save(self.output_file)
        self.log(f"Report saved to: {self.output_file}")

        return str(self.output_file)

    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int) -> None:
        """Apply header styling to a row."""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

    def _auto_width_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # ==========================================================================
    # SHEET 1: DASHBOARD (Executive Summary)
    # ==========================================================================

    def create_dashboard_sheet(self) -> None:
        """
        Create the Dashboard sheet with executive summary and 4 key charts.

        Layout (per PRD):
        - Title and metadata
        - KPI Summary boxes (Max Throughput, Min Latency, Balanced Config)
        - Chart 1: Top 10 Configurations Line Chart
        - Chart 2: Latency vs Throughput Scatter (THE KNEE CHART preview)
        - Chart 3: Scaling Degradation
        - Chart 4: Acks Comparison Summary
        """
        ws = self.wb.create_sheet('Dashboard', 0)
        self.log("Creating Dashboard sheet...")

        # Page setup for printing
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

        # =======================================================================
        # TITLE BANNER
        # =======================================================================
        ws['A1'] = 'Kafka Performance Testing Report'
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].fill = self.HEADER_FILL
        ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 30

        # Metadata row
        test_count = len(self.producer_results) + len(self.consumer_results)
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Tests: {test_count} | Producer: {len(self.producer_results)} | Consumer: {len(self.consumer_results)}'
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:L2')

        # =======================================================================
        # KPI SUMMARY BOXES (Row 4-10)
        # =======================================================================
        self._create_dashboard_kpi_boxes(ws)

        # =======================================================================
        # CHARTS SECTION
        # =======================================================================

        # Chart 1: Top 10 Configurations (A12:G28)
        self._add_top_configs_chart(ws, 'A12')

        # Chart 2: Latency vs Throughput Scatter - KNEE CHART PREVIEW (H12:N28)
        self._add_knee_chart_preview(ws, 'H12')

        # Chart 3: Scaling Degradation (A30:G46)
        self._add_scaling_summary_chart(ws, 'A30')

        # Chart 4: Acks Performance Summary (H30:N46)
        self._add_acks_summary_chart(ws, 'H30')

        # Set column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_dashboard_kpi_boxes(self, ws) -> None:
        """Create KPI summary boxes on dashboard."""
        row = 4

        # Box styling
        kpi_header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        kpi_header_font = Font(bold=True, size=11, color="FFFFFF")

        # =======================================================================
        # KPI BOX 1: MAX THROUGHPUT (A4:C9)
        # =======================================================================
        ws['A4'] = 'MAX THROUGHPUT'
        ws['A4'].fill = kpi_header_fill
        ws['A4'].font = kpi_header_font
        ws.merge_cells('A4:C4')

        if self.stats.get('producer', {}).get('max_throughput_mb'):
            max_mb = self.stats['producer']['max_throughput_mb']
            ws['A5'] = f"{max_mb:.1f} MB/sec"
            ws['A5'].font = Font(bold=True, size=20, color="00B050")
            ws.merge_cells('A5:C5')

            max_rps = self.stats['producer'].get('max_throughput_rps', 0)
            ws['A6'] = f"{max_rps:,.0f} records/sec"
            ws['A6'].font = Font(size=11)
            ws.merge_cells('A6:C6')

            # Best config for throughput
            if self.scores.get('best_throughput_idx') is not None:
                best = self.producer_df.loc[self.scores['best_throughput_idx']]
                ws['A7'] = f"acks={best['acks']}, batch={best['batch_size']}"
                ws['A7'].font = Font(size=9, color="666666")
                ws.merge_cells('A7:C7')
        else:
            ws['A5'] = 'No data'
            ws.merge_cells('A5:C5')

        # =======================================================================
        # KPI BOX 2: MIN LATENCY (D4:F9)
        # =======================================================================
        ws['D4'] = 'MIN LATENCY'
        ws['D4'].fill = kpi_header_fill
        ws['D4'].font = kpi_header_font
        ws.merge_cells('D4:F4')

        if self.stats.get('producer', {}).get('min_latency_ms'):
            min_lat = self.stats['producer']['min_latency_ms']
            ws['D5'] = f"{min_lat:.0f} ms"
            ws['D5'].font = Font(bold=True, size=20, color="FF6B6B")
            ws.merge_cells('D5:F5')

            # Best config for latency
            if self.scores.get('best_latency_idx') is not None:
                best = self.producer_df.loc[self.scores['best_latency_idx']]
                ws['D6'] = f"acks={best['acks']}"
                ws['D6'].font = Font(size=11)
                ws.merge_cells('D6:F6')
                ws['D7'] = f"batch={best['batch_size']}"
                ws['D7'].font = Font(size=9, color="666666")
                ws.merge_cells('D7:F7')
        else:
            ws['D5'] = 'No data'
            ws.merge_cells('D5:F5')

        # =======================================================================
        # KPI BOX 3: BALANCED CONFIG (G4:I9)
        # =======================================================================
        ws['G4'] = 'BALANCED CONFIG'
        ws['G4'].fill = kpi_header_fill
        ws['G4'].font = kpi_header_font
        ws.merge_cells('G4:I4')

        if self.scores.get('best_balanced_idx') is not None:
            best = self.scores['scored_df'].loc[self.scores['best_balanced_idx']]
            ws['G5'] = f"{best['throughput_mb']:.1f} MB/sec"
            ws['G5'].font = Font(bold=True, size=14, color="4472C4")
            ws.merge_cells('G5:I5')
            ws['G6'] = f"{best['avg_latency_ms']:.0f} ms latency"
            ws['G6'].font = Font(size=11)
            ws.merge_cells('G6:I6')
            ws['G7'] = f"Score: {best['score_balanced']:.0f}/100"
            ws['G7'].font = Font(size=9, color="666666")
            ws.merge_cells('G7:I7')
        else:
            ws['G5'] = 'No data'
            ws.merge_cells('G5:I5')

        # =======================================================================
        # KPI BOX 4: SCALING IMPACT (J4:L9)
        # =======================================================================
        ws['J4'] = 'SCALING IMPACT'
        ws['J4'].fill = kpi_header_fill
        ws['J4'].font = kpi_header_font
        ws.merge_cells('J4:L4')

        # Calculate scaling degradation
        scaling_info = self._calculate_scaling_degradation()
        if scaling_info:
            ws['J5'] = f"{scaling_info['degradation_pct']:.1f}% drop"
            ws['J5'].font = Font(bold=True, size=14, color="FF6B6B" if scaling_info['degradation_pct'] > 30 else "FFD93D")
            ws.merge_cells('J5:L5')
            ws['J6'] = f"1 prod: {scaling_info['single_throughput']:.1f} MB/s"
            ws['J6'].font = Font(size=10)
            ws.merge_cells('J6:L6')
            ws['J7'] = f"{scaling_info['multi_count']} prod: {scaling_info['multi_throughput']:.1f} MB/s"
            ws['J7'].font = Font(size=10, color="666666")
            ws.merge_cells('J7:L7')
        else:
            ws['J5'] = 'N/A'
            ws.merge_cells('J5:L5')

    def _calculate_scaling_degradation(self) -> Optional[Dict[str, Any]]:
        """Calculate scaling degradation metrics for dashboard."""
        if self.producer_df.empty or 'num_producers' not in self.producer_df.columns:
            return None

        df = self.producer_df
        grouped = df.groupby('num_producers')['throughput_mb'].mean()

        if len(grouped) < 2:
            return None

        single = grouped.get(1, grouped.iloc[0])
        multi_count = grouped.index.max()
        multi = grouped.get(multi_count, grouped.iloc[-1])

        # Per-producer throughput for multi
        per_producer_multi = multi / multi_count if multi_count > 0 else multi

        degradation = ((single - per_producer_multi) / single) * 100 if single > 0 else 0

        return {
            'single_throughput': single,
            'multi_throughput': multi,
            'multi_count': multi_count,
            'per_producer_multi': per_producer_multi,
            'degradation_pct': degradation,
        }

    def _add_top_configs_chart(self, ws, position: str) -> None:
        """Add Top 10 Configurations line chart to dashboard."""
        if self.producer_df.empty:
            return

        # Get top 10 by throughput
        top_10 = self.producer_df.nlargest(min(10, len(self.producer_df)), 'throughput_mb')

        # Write data to hidden area (columns P-Q)
        data_start_row = 4
        ws['P3'] = 'Config'
        ws['Q3'] = 'Throughput (MB/sec)'

        for idx, (_, row) in enumerate(top_10.iterrows()):
            data_row = data_start_row + idx
            config_label = f"a={row['acks']}_b={row['batch_size']}"
            ws[f'P{data_row}'] = config_label
            ws[f'Q{data_row}'] = row['throughput_mb']

        # Create chart
        chart = BarChart()
        chart.type = "col"
        chart.title = "Top Configurations by Throughput"
        chart.style = self.CHART_STYLE
        chart.y_axis.title = "MB/sec"
        chart.x_axis.title = "Configuration"

        data = Reference(ws, min_col=17, min_row=3, max_row=data_start_row + len(top_10) - 1)
        categories = Reference(ws, min_col=16, min_row=4, max_row=data_start_row + len(top_10) - 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 14
        chart.height = 10

        ws.add_chart(chart, position)

    def _add_knee_chart_preview(self, ws, position: str) -> None:
        """Add Latency vs Throughput scatter chart preview (THE KNEE CHART)."""
        if self.producer_df.empty:
            return

        # Write data to hidden area (columns R-S)
        ws['R3'] = 'Throughput'
        ws['S3'] = 'Latency'

        for idx, (_, row) in enumerate(self.producer_df.iterrows()):
            data_row = 4 + idx
            ws[f'R{data_row}'] = row['throughput_mb']
            ws[f'S{data_row}'] = row['avg_latency_ms']

        # Create scatter chart
        chart = ScatterChart()
        chart.title = "Latency vs Throughput (Saturation Curve)"
        chart.style = self.CHART_STYLE
        chart.x_axis.title = "Throughput (MB/sec)"
        chart.y_axis.title = "Avg Latency (ms)"

        xvalues = Reference(ws, min_col=18, min_row=3, max_row=3 + len(self.producer_df))
        yvalues = Reference(ws, min_col=19, min_row=3, max_row=3 + len(self.producer_df))

        chart.add_data(yvalues, titles_from_data=True)
        chart.set_categories(xvalues)

        if chart.series:
            chart.series[0].marker = Marker(symbol='circle', size=7)
            chart.series[0].graphicalProperties.line.noFill = True
            # Add trendline
            chart.series[0].trendline = Trendline(trendlineType='poly', order=2)

        chart.width = 14
        chart.height = 10

        ws.add_chart(chart, position)

    def _add_scaling_summary_chart(self, ws, position: str) -> None:
        """Add scaling degradation chart to dashboard."""
        if self.producer_df.empty or 'num_producers' not in self.producer_df.columns:
            return

        # Group by num_producers
        scaling = self.producer_df.groupby('num_producers').agg({
            'throughput_mb': 'mean',
            'avg_latency_ms': 'mean'
        }).reset_index()

        if len(scaling) < 2:
            return

        # Write data (columns T-V)
        ws['T3'] = 'Producers'
        ws['U3'] = 'Throughput'
        ws['V3'] = 'Latency'

        for idx, row in scaling.iterrows():
            data_row = 4 + idx
            ws[f'T{data_row}'] = row['num_producers']
            ws[f'U{data_row}'] = row['throughput_mb']
            ws[f'V{data_row}'] = row['avg_latency_ms']

        # Create line chart
        chart = LineChart()
        chart.title = "Scaling Performance"
        chart.style = self.CHART_STYLE
        chart.y_axis.title = "Throughput (MB/sec)"
        chart.x_axis.title = "Number of Producers"

        data = Reference(ws, min_col=21, min_row=3, max_row=3 + len(scaling))
        categories = Reference(ws, min_col=20, min_row=4, max_row=3 + len(scaling))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 14
        chart.height = 10

        ws.add_chart(chart, position)

    def _add_acks_summary_chart(self, ws, position: str) -> None:
        """Add acks setting performance summary chart."""
        if self.producer_df.empty:
            return

        # Group by acks
        acks_data = self.producer_df.groupby('acks').agg({
            'throughput_mb': 'mean',
            'avg_latency_ms': 'mean'
        }).reset_index()

        if acks_data.empty:
            return

        # Write data (columns W-Y)
        ws['W3'] = 'acks'
        ws['X3'] = 'Throughput'
        ws['Y3'] = 'Latency'

        for idx, row in acks_data.iterrows():
            data_row = 4 + idx
            ws[f'W{data_row}'] = str(row['acks'])
            ws[f'X{data_row}'] = row['throughput_mb']
            ws[f'Y{data_row}'] = row['avg_latency_ms']

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.title = "Performance by acks Setting"
        chart.style = self.CHART_STYLE
        chart.y_axis.title = "MB/sec"

        data = Reference(ws, min_col=24, min_row=3, max_row=3 + len(acks_data))
        categories = Reference(ws, min_col=23, min_row=4, max_row=3 + len(acks_data))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 14
        chart.height = 10

        ws.add_chart(chart, position)

    # ==========================================================================
    # SHEET 2: THROUGHPUT ANALYSIS
    # ==========================================================================

    def create_throughput_analysis_sheet(self) -> None:
        """
        Create comprehensive Throughput Analysis sheet with multiple charts.

        Charts included (per PRD):
        1. Throughput by acks Setting (clustered bar)
        2. Throughput by Batch Size (line chart with compression)
        3. Records/Sec vs MB/Sec (scatter)
        4. Compression Efficiency comparison
        """
        ws = self.wb.create_sheet('Throughput Analysis')
        self.log("Creating Throughput Analysis sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'THROUGHPUT OPTIMIZATION ANALYSIS'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'Deep dive into throughput metrics across all test configurations'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # DATA TABLE: All producer results
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'All Producer Test Results'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        # Write headers
        headers = ['Scenario', 'acks', 'batch_size', 'linger_ms', 'compression',
                   'record_size', 'throughput_rps', 'throughput_mb', 'avg_latency_ms']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row += 1

        data_start_row = row
        for _, df_row in self.producer_df.iterrows():
            ws.cell(row=row, column=1, value=df_row['scenario'][:30])  # Truncate long names
            ws.cell(row=row, column=2, value=df_row['acks'])
            ws.cell(row=row, column=3, value=df_row['batch_size'])
            ws.cell(row=row, column=4, value=df_row['linger_ms'])
            ws.cell(row=row, column=5, value=df_row['compression_type'])
            ws.cell(row=row, column=6, value=df_row['record_size'])
            ws.cell(row=row, column=7, value=round(df_row['throughput_rps'], 2))
            ws.cell(row=row, column=8, value=round(df_row['throughput_mb'], 2))
            ws.cell(row=row, column=9, value=round(df_row['avg_latency_ms'], 2))
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.BORDER
            row += 1

        data_end_row = row - 1
        chart_start_row = row + 2

        # =======================================================================
        # CHART 1: Throughput by acks Setting (Clustered Bar)
        # =======================================================================
        acks_grouped = self.producer_df.groupby('acks')['throughput_mb'].mean().reset_index()
        acks_grouped = acks_grouped.sort_values('throughput_mb', ascending=False)

        # Write chart data
        ws[f'K{chart_start_row}'] = 'acks'
        ws[f'L{chart_start_row}'] = 'Avg Throughput'
        for idx, r in acks_grouped.iterrows():
            ws[f'K{chart_start_row + 1 + idx}'] = str(r['acks'])
            ws[f'L{chart_start_row + 1 + idx}'] = r['throughput_mb']

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Throughput by acks Setting"
        chart1.style = self.CHART_STYLE
        chart1.y_axis.title = "MB/sec"
        chart1.x_axis.title = "acks value"

        data = Reference(ws, min_col=12, min_row=chart_start_row,
                        max_row=chart_start_row + len(acks_grouped))
        categories = Reference(ws, min_col=11, min_row=chart_start_row + 1,
                              max_row=chart_start_row + len(acks_grouped))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.width = 12
        chart1.height = 8

        ws.add_chart(chart1, f'A{chart_start_row}')

        # =======================================================================
        # CHART 2: Throughput by Batch Size (Line Chart)
        # =======================================================================
        batch_grouped = self.producer_df.groupby('batch_size')['throughput_mb'].mean().reset_index()
        batch_grouped = batch_grouped.sort_values('batch_size')

        # Write chart data
        batch_data_col = 14
        ws.cell(row=chart_start_row, column=batch_data_col, value='batch_size')
        ws.cell(row=chart_start_row, column=batch_data_col + 1, value='Throughput')

        for idx, r in batch_grouped.iterrows():
            ws.cell(row=chart_start_row + 1 + idx, column=batch_data_col, value=r['batch_size'])
            ws.cell(row=chart_start_row + 1 + idx, column=batch_data_col + 1, value=r['throughput_mb'])

        if len(batch_grouped) > 1:
            chart2 = LineChart()
            chart2.title = "Throughput by Batch Size"
            chart2.style = self.CHART_STYLE
            chart2.y_axis.title = "MB/sec"
            chart2.x_axis.title = "batch.size (bytes)"

            data = Reference(ws, min_col=batch_data_col + 1, min_row=chart_start_row,
                            max_row=chart_start_row + len(batch_grouped))
            categories = Reference(ws, min_col=batch_data_col, min_row=chart_start_row + 1,
                                  max_row=chart_start_row + len(batch_grouped))
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(categories)
            chart2.width = 12
            chart2.height = 8

            ws.add_chart(chart2, f'G{chart_start_row}')

        # =======================================================================
        # CHART 3: Records/Sec vs MB/Sec (Scatter)
        # =======================================================================
        chart3_row = chart_start_row + 18

        # Write chart data
        ws.cell(row=chart3_row, column=11, value='Records/sec')
        ws.cell(row=chart3_row, column=12, value='MB/sec')

        for idx, (_, r) in enumerate(self.producer_df.iterrows()):
            ws.cell(row=chart3_row + 1 + idx, column=11, value=r['throughput_rps'])
            ws.cell(row=chart3_row + 1 + idx, column=12, value=r['throughput_mb'])

        chart3 = ScatterChart()
        chart3.title = "Records/Sec vs MB/Sec Correlation"
        chart3.style = self.CHART_STYLE
        chart3.x_axis.title = "Records/sec"
        chart3.y_axis.title = "MB/sec"

        xvalues = Reference(ws, min_col=11, min_row=chart3_row,
                           max_row=chart3_row + len(self.producer_df))
        yvalues = Reference(ws, min_col=12, min_row=chart3_row,
                           max_row=chart3_row + len(self.producer_df))
        chart3.add_data(yvalues, titles_from_data=True)
        chart3.set_categories(xvalues)
        if chart3.series:
            chart3.series[0].marker = Marker(symbol='circle', size=7)
            chart3.series[0].graphicalProperties.line.noFill = True
        chart3.width = 12
        chart3.height = 8

        ws.add_chart(chart3, f'A{chart3_row}')

        # =======================================================================
        # CHART 4: Compression Efficiency
        # =======================================================================
        compression_grouped = self.producer_df.groupby('compression_type')['throughput_mb'].mean().reset_index()

        # Write chart data
        ws.cell(row=chart3_row, column=14, value='Compression')
        ws.cell(row=chart3_row, column=15, value='Throughput')

        for idx, r in compression_grouped.iterrows():
            ws.cell(row=chart3_row + 1 + idx, column=14, value=r['compression_type'])
            ws.cell(row=chart3_row + 1 + idx, column=15, value=r['throughput_mb'])

        if len(compression_grouped) > 1:
            chart4 = BarChart()
            chart4.type = "col"
            chart4.title = "Compression Type Impact"
            chart4.style = self.CHART_STYLE
            chart4.y_axis.title = "MB/sec"

            data = Reference(ws, min_col=15, min_row=chart3_row,
                            max_row=chart3_row + len(compression_grouped))
            categories = Reference(ws, min_col=14, min_row=chart3_row + 1,
                                  max_row=chart3_row + len(compression_grouped))
            chart4.add_data(data, titles_from_data=True)
            chart4.set_categories(categories)
            chart4.width = 12
            chart4.height = 8

            ws.add_chart(chart4, f'G{chart3_row}')

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 3: LATENCY ANALYSIS
    # ==========================================================================

    def create_latency_analysis_sheet(self) -> None:
        """
        Create comprehensive Latency Analysis sheet.

        Charts included (per PRD):
        1. Latency Percentile Distribution (column chart)
        2. Latency by acks Setting (grouped bars)
        3. Latency Trend across configurations
        4. Latency vs Batch Size
        """
        ws = self.wb.create_sheet('Latency Analysis')
        self.log("Creating Latency Analysis sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'LATENCY ANALYSIS'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'Percentile distributions and latency trends'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # DATA TABLE: Latency metrics
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Latency Metrics by Configuration'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        headers = ['Config', 'acks', 'batch_size', 'avg_ms', 'p50_ms', 'p95_ms', 'p99_ms', 'p999_ms', 'max_ms']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row += 1

        for _, df_row in self.producer_df.iterrows():
            config_label = f"a={df_row['acks']}_b={df_row['batch_size']}"
            ws.cell(row=row, column=1, value=config_label)
            ws.cell(row=row, column=2, value=df_row['acks'])
            ws.cell(row=row, column=3, value=df_row['batch_size'])
            ws.cell(row=row, column=4, value=round(df_row['avg_latency_ms'], 1))
            ws.cell(row=row, column=5, value=df_row['p50_ms'])
            ws.cell(row=row, column=6, value=df_row['p95_ms'])
            ws.cell(row=row, column=7, value=df_row['p99_ms'])
            ws.cell(row=row, column=8, value=df_row['p999_ms'])
            ws.cell(row=row, column=9, value=round(df_row['max_latency_ms'], 1))
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.BORDER
            row += 1

        chart_start_row = row + 2

        # =======================================================================
        # CHART 1: Percentile Distribution (for best and worst configs)
        # =======================================================================
        # Get best (lowest latency) and worst (highest latency) configs
        best_idx = self.producer_df['avg_latency_ms'].idxmin()
        worst_idx = self.producer_df['avg_latency_ms'].idxmax()
        best = self.producer_df.loc[best_idx]
        worst = self.producer_df.loc[worst_idx]

        # Write percentile data
        percentiles = ['p50', 'p95', 'p99', 'p999']
        ws.cell(row=chart_start_row, column=11, value='Percentile')
        ws.cell(row=chart_start_row, column=12, value='Best Config')
        ws.cell(row=chart_start_row, column=13, value='Worst Config')

        for idx, pct in enumerate(percentiles):
            ws.cell(row=chart_start_row + 1 + idx, column=11, value=pct)
            ws.cell(row=chart_start_row + 1 + idx, column=12, value=best[f'{pct}_ms'])
            ws.cell(row=chart_start_row + 1 + idx, column=13, value=worst[f'{pct}_ms'])

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Latency Percentile Distribution"
        chart1.style = self.CHART_STYLE
        chart1.y_axis.title = "Latency (ms)"
        chart1.x_axis.title = "Percentile"

        data = Reference(ws, min_col=12, max_col=13, min_row=chart_start_row,
                        max_row=chart_start_row + len(percentiles))
        categories = Reference(ws, min_col=11, min_row=chart_start_row + 1,
                              max_row=chart_start_row + len(percentiles))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.width = 12
        chart1.height = 8

        ws.add_chart(chart1, f'A{chart_start_row}')

        # =======================================================================
        # CHART 2: Latency by acks Setting
        # =======================================================================
        acks_latency = self.producer_df.groupby('acks').agg({
            'avg_latency_ms': 'mean',
            'p99_ms': 'mean'
        }).reset_index()

        ws.cell(row=chart_start_row, column=15, value='acks')
        ws.cell(row=chart_start_row, column=16, value='Avg Latency')
        ws.cell(row=chart_start_row, column=17, value='P99 Latency')

        for idx, r in acks_latency.iterrows():
            ws.cell(row=chart_start_row + 1 + idx, column=15, value=str(r['acks']))
            ws.cell(row=chart_start_row + 1 + idx, column=16, value=round(r['avg_latency_ms'], 1))
            ws.cell(row=chart_start_row + 1 + idx, column=17, value=round(r['p99_ms'], 1))

        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Latency by acks Setting"
        chart2.style = self.CHART_STYLE
        chart2.y_axis.title = "Latency (ms)"

        data = Reference(ws, min_col=16, max_col=17, min_row=chart_start_row,
                        max_row=chart_start_row + len(acks_latency))
        categories = Reference(ws, min_col=15, min_row=chart_start_row + 1,
                              max_row=chart_start_row + len(acks_latency))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(categories)
        chart2.width = 12
        chart2.height = 8

        ws.add_chart(chart2, f'G{chart_start_row}')

        # =======================================================================
        # CHART 3: Latency Trend Line (Avg, P95, P99)
        # =======================================================================
        chart3_row = chart_start_row + 18

        ws.cell(row=chart3_row, column=11, value='Config')
        ws.cell(row=chart3_row, column=12, value='Avg')
        ws.cell(row=chart3_row, column=13, value='P95')
        ws.cell(row=chart3_row, column=14, value='P99')

        for idx, (_, r) in enumerate(self.producer_df.iterrows()):
            config_label = f"a={r['acks']}"
            ws.cell(row=chart3_row + 1 + idx, column=11, value=config_label)
            ws.cell(row=chart3_row + 1 + idx, column=12, value=round(r['avg_latency_ms'], 1))
            ws.cell(row=chart3_row + 1 + idx, column=13, value=r['p95_ms'])
            ws.cell(row=chart3_row + 1 + idx, column=14, value=r['p99_ms'])

        chart3 = LineChart()
        chart3.title = "Latency Trend (Avg, P95, P99)"
        chart3.style = self.CHART_STYLE
        chart3.y_axis.title = "Latency (ms)"

        data = Reference(ws, min_col=12, max_col=14, min_row=chart3_row,
                        max_row=chart3_row + len(self.producer_df))
        categories = Reference(ws, min_col=11, min_row=chart3_row + 1,
                              max_row=chart3_row + len(self.producer_df))
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(categories)
        chart3.width = 12
        chart3.height = 8

        ws.add_chart(chart3, f'A{chart3_row}')

        # =======================================================================
        # CHART 4: Latency vs Batch Size (Scatter)
        # =======================================================================
        ws.cell(row=chart3_row, column=16, value='Batch Size')
        ws.cell(row=chart3_row, column=17, value='Avg Latency')

        for idx, (_, r) in enumerate(self.producer_df.iterrows()):
            ws.cell(row=chart3_row + 1 + idx, column=16, value=r['batch_size'])
            ws.cell(row=chart3_row + 1 + idx, column=17, value=r['avg_latency_ms'])

        chart4 = ScatterChart()
        chart4.title = "Latency vs Batch Size"
        chart4.style = self.CHART_STYLE
        chart4.x_axis.title = "Batch Size (bytes)"
        chart4.y_axis.title = "Avg Latency (ms)"

        xvalues = Reference(ws, min_col=16, min_row=chart3_row,
                           max_row=chart3_row + len(self.producer_df))
        yvalues = Reference(ws, min_col=17, min_row=chart3_row,
                           max_row=chart3_row + len(self.producer_df))
        chart4.add_data(yvalues, titles_from_data=True)
        chart4.set_categories(xvalues)
        if chart4.series:
            chart4.series[0].marker = Marker(symbol='circle', size=7)
            chart4.series[0].graphicalProperties.line.noFill = True
        chart4.width = 12
        chart4.height = 8

        ws.add_chart(chart4, f'G{chart3_row}')

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 4: TRADE-OFF ANALYSIS (THE CRITICAL KNEE CHART)
    # ==========================================================================

    def create_tradeoff_analysis_sheet(self) -> None:
        """
        Create the Trade-off Analysis sheet with THE CRITICAL KNEE CHART.

        This is the MOST IMPORTANT visualization showing:
        - Latency vs Throughput scatter plot with saturation zones
        - The "knee" point where latency explodes
        - Color-coded by acks setting
        - Performance scoring breakdown
        """
        ws = self.wb.create_sheet('Trade-off Analysis')
        self.log("Creating Trade-off Analysis sheet (THE KNEE CHART)...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'TRADE-OFF ANALYSIS: The Saturation Curve'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:L1')

        ws['A2'] = 'THE CRITICAL CHART: Shows where system saturates (latency explodes)'
        ws['A2'].font = Font(bold=True, italic=True, color="C00000")
        ws.merge_cells('A2:L2')

        # =======================================================================
        # ZONE LEGEND
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Performance Zones:'
        ws[f'A{row}'].font = self.SECTION_FONT

        ws[f'B{row}'] = 'Safe Zone (0-50 MB/s, <1000ms)'
        ws[f'B{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        ws[f'D{row}'] = 'Diminishing Returns (50-90 MB/s)'
        ws[f'D{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws[f'F{row}'] = 'Saturation (>90 MB/s, high latency)'
        ws[f'F{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # =======================================================================
        # DATA TABLE with performance zones
        # =======================================================================
        row = 6
        ws[f'A{row}'] = 'Configuration Performance Data'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        headers = ['Config', 'acks', 'Throughput (MB/s)', 'Latency (ms)', 'P99 (ms)', 'Zone', 'Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row += 1

        scored_df = self.scores.get('scored_df', self.producer_df)

        for _, r in scored_df.iterrows():
            config_label = f"a={r['acks']}_b={r['batch_size']}"
            throughput = r['throughput_mb']
            latency = r['avg_latency_ms']

            # Determine zone
            if throughput < 50 and latency < 1000:
                zone = 'Safe'
                zone_fill = self.GOOD_FILL
            elif throughput < 90:
                zone = 'Diminishing'
                zone_fill = self.NEUTRAL_FILL
            else:
                zone = 'Saturation'
                zone_fill = self.BAD_FILL

            ws.cell(row=row, column=1, value=config_label)
            ws.cell(row=row, column=2, value=r['acks'])
            ws.cell(row=row, column=3, value=round(throughput, 2))
            ws.cell(row=row, column=4, value=round(latency, 1))
            ws.cell(row=row, column=5, value=r['p99_ms'])
            ws.cell(row=row, column=6, value=zone)
            ws.cell(row=row, column=6).fill = zone_fill

            # Score
            score = r.get('score_balanced', 0)
            ws.cell(row=row, column=7, value=round(score, 1))

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.BORDER
            row += 1

        chart_start_row = row + 2

        # =======================================================================
        # THE KNEE CHART: Latency vs Throughput (MAIN VISUALIZATION)
        # =======================================================================
        # Write chart data with acks grouping
        ws.cell(row=chart_start_row, column=10, value='Throughput')
        ws.cell(row=chart_start_row, column=11, value='Latency')
        ws.cell(row=chart_start_row, column=12, value='acks')

        for idx, (_, r) in enumerate(self.producer_df.iterrows()):
            ws.cell(row=chart_start_row + 1 + idx, column=10, value=r['throughput_mb'])
            ws.cell(row=chart_start_row + 1 + idx, column=11, value=r['avg_latency_ms'])
            ws.cell(row=chart_start_row + 1 + idx, column=12, value=r['acks'])

        # Create THE KNEE CHART
        chart = ScatterChart()
        chart.title = "THE KNEE CHART: Latency vs Throughput Saturation"
        chart.style = self.CHART_STYLE
        chart.x_axis.title = "Throughput (MB/sec)"
        chart.y_axis.title = "Average Latency (ms)"
        chart.x_axis.scaling.min = 0
        chart.y_axis.scaling.min = 0

        # Create series for all data points
        xvalues = Reference(ws, min_col=10, min_row=chart_start_row,
                           max_row=chart_start_row + len(self.producer_df))
        yvalues = Reference(ws, min_col=11, min_row=chart_start_row,
                           max_row=chart_start_row + len(self.producer_df))

        chart.add_data(yvalues, titles_from_data=True)
        chart.set_categories(xvalues)

        if chart.series:
            chart.series[0].marker = Marker(symbol='circle', size=10)
            chart.series[0].graphicalProperties.line.noFill = True
            # Add polynomial trendline to show the "knee"
            chart.series[0].trendline = Trendline(trendlineType='poly', order=2, dispEq=False, dispRSqr=False)

        chart.width = 16
        chart.height = 12

        ws.add_chart(chart, f'A{chart_start_row}')

        # =======================================================================
        # PERFORMANCE SCORE BREAKDOWN CHART
        # =======================================================================
        score_chart_row = chart_start_row + 26

        ws[f'A{score_chart_row}'] = 'Performance Score Breakdown'
        ws[f'A{score_chart_row}'].font = self.SECTION_FONT

        # Write score data
        ws.cell(row=score_chart_row + 1, column=10, value='Config')
        ws.cell(row=score_chart_row + 1, column=11, value='Throughput Score')
        ws.cell(row=score_chart_row + 1, column=12, value='Latency Score')
        ws.cell(row=score_chart_row + 1, column=13, value='Consistency Score')

        top_5 = scored_df.nlargest(5, 'score_balanced')
        for idx, (_, r) in enumerate(top_5.iterrows()):
            config_label = f"a={r['acks']}"
            ws.cell(row=score_chart_row + 2 + idx, column=10, value=config_label)
            ws.cell(row=score_chart_row + 2 + idx, column=11, value=round(r['throughput_score'], 1))
            ws.cell(row=score_chart_row + 2 + idx, column=12, value=round(r['latency_score'], 1))
            ws.cell(row=score_chart_row + 2 + idx, column=13, value=round(r['consistency_score'], 1))

        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Top 5 Configurations - Score Breakdown"
        chart2.style = self.CHART_STYLE
        chart2.y_axis.title = "Score (0-100)"

        data = Reference(ws, min_col=11, max_col=13, min_row=score_chart_row + 1,
                        max_row=score_chart_row + 1 + len(top_5))
        categories = Reference(ws, min_col=10, min_row=score_chart_row + 2,
                              max_row=score_chart_row + 1 + len(top_5))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(categories)
        chart2.width = 14
        chart2.height = 8

        ws.add_chart(chart2, f'A{score_chart_row + 2}')

        # =======================================================================
        # SCORING METHODOLOGY EXPLANATION
        # =======================================================================
        method_row = score_chart_row + 2
        ws.cell(row=method_row, column=1, value='Scoring Methodology:')
        ws.cell(row=method_row, column=1).font = Font(bold=True)
        ws.cell(row=method_row + 1, column=1, value='Throughput Score = (test_throughput / max_observed) * 100')
        ws.cell(row=method_row + 2, column=1, value='Latency Score = 100 - ((test_latency / max_observed) * 100)')
        ws.cell(row=method_row + 3, column=1, value='Consistency Score = Based on P99/Avg ratio')
        ws.cell(row=method_row + 4, column=1, value='Balanced Score = 30% throughput + 35% latency + 35% consistency')

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 5: SCALING PERFORMANCE
    # ==========================================================================

    def create_scaling_performance_sheet(self) -> None:
        """
        Create Scaling Performance sheet with degradation analysis.

        Charts included (per PRD):
        1. Throughput per Producer line chart (actual vs ideal)
        2. Aggregate Throughput area chart
        3. Latency Impact with scaling
        4. Consumer Scaling analysis
        """
        ws = self.wb.create_sheet('Scaling Performance')
        self.log("Creating Scaling Performance sheet...")

        # Title
        ws['A1'] = 'SCALING PERFORMANCE ANALYSIS'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'How performance degrades with multiple producers/consumers'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # PRODUCER SCALING ANALYSIS
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Producer Scaling Analysis'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        # Group by num_producers
        if not self.producer_df.empty and 'num_producers' in self.producer_df.columns:
            scaling_data = self.producer_df.groupby('num_producers').agg({
                'throughput_mb': ['mean', 'sum'],
                'avg_latency_ms': 'mean'
            }).reset_index()
            scaling_data.columns = ['num_producers', 'avg_throughput', 'total_throughput', 'avg_latency']

            # Calculate per-producer throughput and ideal scaling
            if len(scaling_data) > 0:
                single_throughput = scaling_data[scaling_data['num_producers'] == 1]['avg_throughput'].values
                if len(single_throughput) > 0:
                    single_throughput = single_throughput[0]
                else:
                    single_throughput = scaling_data['avg_throughput'].iloc[0]

                scaling_data['per_producer_throughput'] = scaling_data['total_throughput'] / scaling_data['num_producers']
                scaling_data['ideal_throughput'] = single_throughput
                scaling_data['efficiency_pct'] = (scaling_data['per_producer_throughput'] / single_throughput) * 100

            # Write data table
            headers = ['Num Producers', 'Avg Throughput', 'Total Throughput', 'Per-Producer', 'Ideal', 'Efficiency %', 'Avg Latency']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.BORDER
            row += 1

            data_start_row = row
            for _, r in scaling_data.iterrows():
                ws.cell(row=row, column=1, value=int(r['num_producers']))
                ws.cell(row=row, column=2, value=round(r['avg_throughput'], 2))
                ws.cell(row=row, column=3, value=round(r['total_throughput'], 2))
                ws.cell(row=row, column=4, value=round(r['per_producer_throughput'], 2))
                ws.cell(row=row, column=5, value=round(r['ideal_throughput'], 2))
                ws.cell(row=row, column=6, value=round(r['efficiency_pct'], 1))
                ws.cell(row=row, column=7, value=round(r['avg_latency'], 1))
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = self.BORDER
                row += 1

            chart_start_row = row + 2

            # =======================================================================
            # CHART 1: Throughput per Producer (Actual vs Ideal)
            # =======================================================================
            if len(scaling_data) > 1:
                # Write chart data
                ws.cell(row=chart_start_row, column=10, value='Producers')
                ws.cell(row=chart_start_row, column=11, value='Actual Per-Producer')
                ws.cell(row=chart_start_row, column=12, value='Ideal (Linear)')

                for idx, r in scaling_data.iterrows():
                    ws.cell(row=chart_start_row + 1 + idx, column=10, value=int(r['num_producers']))
                    ws.cell(row=chart_start_row + 1 + idx, column=11, value=round(r['per_producer_throughput'], 2))
                    ws.cell(row=chart_start_row + 1 + idx, column=12, value=round(r['ideal_throughput'], 2))

                chart1 = LineChart()
                chart1.title = "Throughput per Producer: Actual vs Ideal"
                chart1.style = self.CHART_STYLE
                chart1.y_axis.title = "MB/sec per Producer"
                chart1.x_axis.title = "Number of Producers"

                data = Reference(ws, min_col=11, max_col=12, min_row=chart_start_row,
                                max_row=chart_start_row + len(scaling_data))
                categories = Reference(ws, min_col=10, min_row=chart_start_row + 1,
                                      max_row=chart_start_row + len(scaling_data))
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(categories)
                chart1.width = 12
                chart1.height = 8

                ws.add_chart(chart1, f'A{chart_start_row}')

                # =======================================================================
                # CHART 2: Aggregate Throughput (Area Chart)
                # =======================================================================
                ws.cell(row=chart_start_row, column=14, value='Producers')
                ws.cell(row=chart_start_row, column=15, value='Total Throughput')

                for idx, r in scaling_data.iterrows():
                    ws.cell(row=chart_start_row + 1 + idx, column=14, value=int(r['num_producers']))
                    ws.cell(row=chart_start_row + 1 + idx, column=15, value=round(r['total_throughput'], 2))

                chart2 = AreaChart()
                chart2.title = "Aggregate Throughput by Producer Count"
                chart2.style = self.CHART_STYLE
                chart2.y_axis.title = "Total MB/sec"
                chart2.x_axis.title = "Number of Producers"

                data = Reference(ws, min_col=15, min_row=chart_start_row,
                                max_row=chart_start_row + len(scaling_data))
                categories = Reference(ws, min_col=14, min_row=chart_start_row + 1,
                                      max_row=chart_start_row + len(scaling_data))
                chart2.add_data(data, titles_from_data=True)
                chart2.set_categories(categories)
                chart2.width = 12
                chart2.height = 8

                ws.add_chart(chart2, f'G{chart_start_row}')

                # =======================================================================
                # CHART 3: Latency Impact with Scaling
                # =======================================================================
                chart3_row = chart_start_row + 18

                ws.cell(row=chart3_row, column=10, value='Producers')
                ws.cell(row=chart3_row, column=11, value='Avg Latency (ms)')

                for idx, r in scaling_data.iterrows():
                    ws.cell(row=chart3_row + 1 + idx, column=10, value=int(r['num_producers']))
                    ws.cell(row=chart3_row + 1 + idx, column=11, value=round(r['avg_latency'], 1))

                chart3 = BarChart()
                chart3.type = "col"
                chart3.title = "Latency Impact with Scaling"
                chart3.style = self.CHART_STYLE
                chart3.y_axis.title = "Latency (ms)"

                data = Reference(ws, min_col=11, min_row=chart3_row,
                                max_row=chart3_row + len(scaling_data))
                categories = Reference(ws, min_col=10, min_row=chart3_row + 1,
                                      max_row=chart3_row + len(scaling_data))
                chart3.add_data(data, titles_from_data=True)
                chart3.set_categories(categories)
                chart3.width = 12
                chart3.height = 8

                ws.add_chart(chart3, f'A{chart3_row}')

                row = chart3_row + 18
        else:
            ws[f'A{row}'] = 'No multi-producer scaling data available'
            row += 2

        # =======================================================================
        # CONSUMER SCALING ANALYSIS
        # =======================================================================
        ws[f'A{row}'] = 'Consumer Scaling Analysis'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        if not self.consumer_df.empty and 'num_consumers' in self.consumer_df.columns:
            consumer_scaling = self.consumer_df.groupby('num_consumers').agg({
                'throughput_mb_sec': 'mean',
                'rebalance_time_ms': 'mean'
            }).reset_index()

            # Write data table
            headers = ['Num Consumers', 'Avg Throughput (MB/sec)', 'Avg Rebalance Time (ms)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.BORDER
            row += 1

            for _, r in consumer_scaling.iterrows():
                ws.cell(row=row, column=1, value=int(r['num_consumers']))
                ws.cell(row=row, column=2, value=round(r['throughput_mb_sec'], 2))
                ws.cell(row=row, column=3, value=round(r['rebalance_time_ms'], 0))
                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = self.BORDER
                row += 1

            # Consumer throughput chart
            if len(consumer_scaling) > 1:
                chart_row = row + 2

                ws.cell(row=chart_row, column=10, value='Consumers')
                ws.cell(row=chart_row, column=11, value='Throughput')

                for idx, r in consumer_scaling.iterrows():
                    ws.cell(row=chart_row + 1 + idx, column=10, value=int(r['num_consumers']))
                    ws.cell(row=chart_row + 1 + idx, column=11, value=round(r['throughput_mb_sec'], 2))

                chart = BarChart()
                chart.type = "col"
                chart.title = "Consumer Throughput by Scale"
                chart.style = self.CHART_STYLE
                chart.y_axis.title = "MB/sec"

                data = Reference(ws, min_col=11, min_row=chart_row,
                                max_row=chart_row + len(consumer_scaling))
                categories = Reference(ws, min_col=10, min_row=chart_row + 1,
                                      max_row=chart_row + len(consumer_scaling))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.width = 12
                chart.height = 8

                ws.add_chart(chart, f'G{row + 2}')
        else:
            ws[f'A{row}'] = 'No consumer scaling data available'

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 6: CONFIGURATION HEATMAP
    # ==========================================================================

    def create_heatmap_sheet(self) -> None:
        """
        Create Configuration Heatmap sheet with multi-dimensional analysis.

        Heatmaps included (per PRD):
        1. Throughput Heatmap: batch_size vs acks
        2. Latency Heatmap: acks vs compression
        3. Compression Efficiency by record_size
        """
        ws = self.wb.create_sheet('Configuration Heatmap')
        self.log("Creating Configuration Heatmap sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'CONFIGURATION HEATMAP ANALYSIS'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'Multi-dimensional parameter interaction analysis (colors indicate performance)'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # HEATMAP 1: Throughput by acks and batch_size
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Throughput Heatmap: acks vs batch_size'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        # Create pivot table
        pivot = self.producer_df.pivot_table(
            values='throughput_mb',
            index='acks',
            columns='batch_size',
            aggfunc='mean'
        )

        if not pivot.empty:
            # Write column headers (batch_size values)
            ws.cell(row=row, column=1, value='acks \\ batch_size')
            ws.cell(row=row, column=1).font = self.HEADER_FONT
            ws.cell(row=row, column=1).fill = self.HEADER_FILL
            ws.cell(row=row, column=1).border = self.BORDER

            for col_idx, batch in enumerate(pivot.columns):
                cell = ws.cell(row=row, column=col_idx + 2, value=batch)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.BORDER
            row += 1

            # Get min/max for color scaling
            min_val = pivot.min().min()
            max_val = pivot.max().max()
            val_range = max_val - min_val if max_val > min_val else 1

            # Write data with conditional coloring
            for acks_val in pivot.index:
                ws.cell(row=row, column=1, value=str(acks_val))
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).border = self.BORDER

                for col_idx, batch in enumerate(pivot.columns):
                    value = pivot.loc[acks_val, batch]
                    cell = ws.cell(row=row, column=col_idx + 2, value=round(value, 1) if not pd.isna(value) else '')
                    cell.border = self.BORDER

                    # Color based on value (green = high throughput, red = low)
                    if not pd.isna(value):
                        intensity = int(((value - min_val) / val_range) * 155) + 100
                        # Green for high values
                        cell.fill = PatternFill(
                            start_color=f"{255-intensity:02X}{intensity:02X}50",
                            end_color=f"{255-intensity:02X}{intensity:02X}50",
                            fill_type="solid"
                        )
                row += 1

            row += 2

        # =======================================================================
        # HEATMAP 2: Latency by acks and compression
        # =======================================================================
        ws[f'A{row}'] = 'Latency Heatmap: acks vs compression_type'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        pivot2 = self.producer_df.pivot_table(
            values='avg_latency_ms',
            index='acks',
            columns='compression_type',
            aggfunc='mean'
        )

        if not pivot2.empty:
            # Write column headers
            ws.cell(row=row, column=1, value='acks \\ compression')
            ws.cell(row=row, column=1).font = self.HEADER_FONT
            ws.cell(row=row, column=1).fill = self.HEADER_FILL
            ws.cell(row=row, column=1).border = self.BORDER

            for col_idx, comp in enumerate(pivot2.columns):
                cell = ws.cell(row=row, column=col_idx + 2, value=comp)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.BORDER
            row += 1

            # Get min/max for color scaling
            min_val = pivot2.min().min()
            max_val = pivot2.max().max()
            val_range = max_val - min_val if max_val > min_val else 1

            # Write data with conditional coloring (inverted - green = low latency)
            for acks_val in pivot2.index:
                ws.cell(row=row, column=1, value=str(acks_val))
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).border = self.BORDER

                for col_idx, comp in enumerate(pivot2.columns):
                    value = pivot2.loc[acks_val, comp]
                    cell = ws.cell(row=row, column=col_idx + 2, value=round(value, 0) if not pd.isna(value) else '')
                    cell.border = self.BORDER

                    # Color based on value (green = low latency, red = high)
                    if not pd.isna(value):
                        intensity = int(((max_val - value) / val_range) * 155) + 100
                        cell.fill = PatternFill(
                            start_color=f"{255-intensity:02X}{intensity:02X}50",
                            end_color=f"{255-intensity:02X}{intensity:02X}50",
                            fill_type="solid"
                        )
                row += 1

            row += 2

        # =======================================================================
        # COLOR LEGEND
        # =======================================================================
        ws[f'A{row}'] = 'Color Legend:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = 'Throughput Heatmap: Green = Higher throughput (better)'
        ws[f'A{row}'].fill = PatternFill(start_color="64FF50", end_color="64FF50", fill_type="solid")
        ws[f'C{row}'] = 'Red = Lower throughput'
        ws[f'C{row}'].fill = PatternFill(start_color="FF6450", end_color="FF6450", fill_type="solid")
        row += 1

        ws[f'A{row}'] = 'Latency Heatmap: Green = Lower latency (better)'
        ws[f'A{row}'].fill = PatternFill(start_color="64FF50", end_color="64FF50", fill_type="solid")
        ws[f'C{row}'] = 'Red = Higher latency'
        ws[f'C{row}'].fill = PatternFill(start_color="FF6450", end_color="FF6450", fill_type="solid")
        row += 2

        # =======================================================================
        # SUMMARY INSIGHTS
        # =======================================================================
        ws[f'A{row}'] = 'Key Insights from Heatmaps:'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        # Find optimal combinations
        if not self.producer_df.empty:
            best_throughput = self.producer_df.loc[self.producer_df['throughput_mb'].idxmax()]
            best_latency = self.producer_df.loc[self.producer_df['avg_latency_ms'].idxmin()]

            ws[f'A{row}'] = f"Best Throughput: acks={best_throughput['acks']}, batch={best_throughput['batch_size']}, compression={best_throughput['compression_type']} -> {best_throughput['throughput_mb']:.1f} MB/sec"
            row += 1
            ws[f'A{row}'] = f"Best Latency: acks={best_latency['acks']}, batch={best_latency['batch_size']}, compression={best_latency['compression_type']} -> {best_latency['avg_latency_ms']:.0f} ms"

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 7: MESSAGE SIZE IMPACT
    # ==========================================================================

    def create_message_size_sheet(self) -> None:
        """
        Create Message Size Impact sheet with efficiency analysis.

        Charts included (per PRD):
        1. Throughput vs Message Size
        2. Records/Sec vs Message Size
        3. Compression efficiency by size
        4. Latency stability by size
        """
        ws = self.wb.create_sheet('Message Size Impact')
        self.log("Creating Message Size Impact sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'MESSAGE SIZE IMPACT ANALYSIS'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'How message size affects throughput and efficiency'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # DATA TABLE
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Performance by Record Size'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        # Group by record_size
        size_data = self.producer_df.groupby('record_size').agg({
            'throughput_mb': 'mean',
            'throughput_rps': 'mean',
            'avg_latency_ms': 'mean',
            'p99_ms': 'mean'
        }).reset_index()
        size_data = size_data.sort_values('record_size')

        headers = ['Record Size (bytes)', 'Avg Throughput (MB/sec)', 'Avg Records/sec', 'Avg Latency (ms)', 'P99 Latency (ms)', 'Test Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row += 1

        test_counts = self.producer_df.groupby('record_size').size()
        for _, r in size_data.iterrows():
            ws.cell(row=row, column=1, value=int(r['record_size']))
            ws.cell(row=row, column=2, value=round(r['throughput_mb'], 2))
            ws.cell(row=row, column=3, value=round(r['throughput_rps'], 0))
            ws.cell(row=row, column=4, value=round(r['avg_latency_ms'], 1))
            ws.cell(row=row, column=5, value=round(r['p99_ms'], 1))
            ws.cell(row=row, column=6, value=test_counts.get(r['record_size'], 0))
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.BORDER
            row += 1

        chart_start_row = row + 2

        # =======================================================================
        # CHART 1: Throughput vs Message Size
        # =======================================================================
        if len(size_data) > 1:
            ws.cell(row=chart_start_row, column=10, value='Size')
            ws.cell(row=chart_start_row, column=11, value='Throughput')

            for idx, r in size_data.iterrows():
                ws.cell(row=chart_start_row + 1 + idx, column=10, value=int(r['record_size']))
                ws.cell(row=chart_start_row + 1 + idx, column=11, value=round(r['throughput_mb'], 2))

            chart1 = LineChart()
            chart1.title = "Throughput vs Message Size"
            chart1.style = self.CHART_STYLE
            chart1.y_axis.title = "MB/sec"
            chart1.x_axis.title = "Record Size (bytes)"

            data = Reference(ws, min_col=11, min_row=chart_start_row,
                            max_row=chart_start_row + len(size_data))
            categories = Reference(ws, min_col=10, min_row=chart_start_row + 1,
                                  max_row=chart_start_row + len(size_data))
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(categories)
            chart1.width = 12
            chart1.height = 8

            ws.add_chart(chart1, f'A{chart_start_row}')

            # =======================================================================
            # CHART 2: Records/Sec vs Message Size
            # =======================================================================
            ws.cell(row=chart_start_row, column=13, value='Size')
            ws.cell(row=chart_start_row, column=14, value='Records/sec')

            for idx, r in size_data.iterrows():
                ws.cell(row=chart_start_row + 1 + idx, column=13, value=int(r['record_size']))
                ws.cell(row=chart_start_row + 1 + idx, column=14, value=round(r['throughput_rps'], 0))

            chart2 = LineChart()
            chart2.title = "Records/Sec vs Message Size"
            chart2.style = self.CHART_STYLE
            chart2.y_axis.title = "Records/sec"
            chart2.x_axis.title = "Record Size (bytes)"

            data = Reference(ws, min_col=14, min_row=chart_start_row,
                            max_row=chart_start_row + len(size_data))
            categories = Reference(ws, min_col=13, min_row=chart_start_row + 1,
                                  max_row=chart_start_row + len(size_data))
            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(categories)
            chart2.width = 12
            chart2.height = 8

            ws.add_chart(chart2, f'G{chart_start_row}')

            # =======================================================================
            # CHART 3: Latency vs Message Size
            # =======================================================================
            chart3_row = chart_start_row + 18

            ws.cell(row=chart3_row, column=10, value='Size')
            ws.cell(row=chart3_row, column=11, value='Avg Latency')
            ws.cell(row=chart3_row, column=12, value='P99 Latency')

            for idx, r in size_data.iterrows():
                ws.cell(row=chart3_row + 1 + idx, column=10, value=int(r['record_size']))
                ws.cell(row=chart3_row + 1 + idx, column=11, value=round(r['avg_latency_ms'], 1))
                ws.cell(row=chart3_row + 1 + idx, column=12, value=round(r['p99_ms'], 1))

            chart3 = LineChart()
            chart3.title = "Latency vs Message Size"
            chart3.style = self.CHART_STYLE
            chart3.y_axis.title = "Latency (ms)"
            chart3.x_axis.title = "Record Size (bytes)"

            data = Reference(ws, min_col=11, max_col=12, min_row=chart3_row,
                            max_row=chart3_row + len(size_data))
            categories = Reference(ws, min_col=10, min_row=chart3_row + 1,
                                  max_row=chart3_row + len(size_data))
            chart3.add_data(data, titles_from_data=True)
            chart3.set_categories(categories)
            chart3.width = 12
            chart3.height = 8

            ws.add_chart(chart3, f'A{chart3_row}')

        # =======================================================================
        # INSIGHTS
        # =======================================================================
        insight_row = row + 40 if len(size_data) > 1 else row + 2
        ws[f'A{insight_row}'] = 'Key Insights:'
        ws[f'A{insight_row}'].font = self.SECTION_FONT
        insight_row += 1

        if not size_data.empty:
            best_throughput_size = size_data.loc[size_data['throughput_mb'].idxmax(), 'record_size']
            best_rps_size = size_data.loc[size_data['throughput_rps'].idxmax(), 'record_size']
            ws[f'A{insight_row}'] = f"Best MB/sec at record size: {int(best_throughput_size)} bytes"
            insight_row += 1
            ws[f'A{insight_row}'] = f"Best records/sec at record size: {int(best_rps_size)} bytes"
            insight_row += 1
            ws[f'A{insight_row}'] = "Trade-off: Larger messages = higher MB/sec but lower records/sec"

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 8: ACKS COMPARISON
    # ==========================================================================

    def create_acks_comparison_sheet(self) -> None:
        """
        Create Acks Comparison sheet with durability vs performance trade-off.

        Charts included (per PRD):
        1. Throughput by acks setting (bar)
        2. Latency by acks setting (bar)
        3. Performance vs Risk trade-off (scatter)
        4. Detailed comparison table
        """
        ws = self.wb.create_sheet('Acks Comparison')
        self.log("Creating Acks Comparison sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No producer test results available'
            return

        # Title
        ws['A1'] = 'DURABILITY vs PERFORMANCE: acks Setting Analysis'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'Trade-off analysis between data safety and throughput'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        # =======================================================================
        # ACKS EXPLANATION
        # =======================================================================
        row = 4
        ws[f'A{row}'] = 'Understanding acks Settings:'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        explanations = [
            ('acks=0', 'Fire-and-forget. No acknowledgment. FASTEST but messages can be lost.', 'FF6B6B'),
            ('acks=1', 'Leader only. Waits for leader to write. BALANCED speed/safety.', 'FFD93D'),
            ('acks=all/-1', 'All replicas. Waits for all in-sync replicas. SAFEST but slowest.', '6BCB77'),
        ]

        for acks, desc, color in explanations:
            ws[f'A{row}'] = acks
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f'B{row}'] = desc
            ws.merge_cells(f'B{row}:G{row}')
            row += 1

        row += 1

        # =======================================================================
        # DATA TABLE
        # =======================================================================
        ws[f'A{row}'] = 'Performance by acks Setting'
        ws[f'A{row}'].font = self.SECTION_FONT
        row += 1

        acks_data = self.producer_df.groupby('acks').agg({
            'throughput_mb': ['mean', 'max'],
            'avg_latency_ms': ['mean', 'min'],
            'p99_ms': 'mean'
        }).reset_index()
        acks_data.columns = ['acks', 'avg_throughput', 'max_throughput', 'avg_latency', 'min_latency', 'p99_latency']

        headers = ['acks', 'Avg Throughput', 'Max Throughput', 'Avg Latency', 'Min Latency', 'P99 Latency', 'Test Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER
        row += 1

        test_counts = self.producer_df.groupby('acks').size()
        for _, r in acks_data.iterrows():
            acks_val = str(r['acks'])
            ws.cell(row=row, column=1, value=acks_val)

            # Color code by acks value
            acks_color = self.ACKS_COLORS.get(acks_val, 'FFFFFF')
            ws.cell(row=row, column=1).fill = PatternFill(start_color=acks_color, end_color=acks_color, fill_type="solid")

            ws.cell(row=row, column=2, value=round(r['avg_throughput'], 2))
            ws.cell(row=row, column=3, value=round(r['max_throughput'], 2))
            ws.cell(row=row, column=4, value=round(r['avg_latency'], 1))
            ws.cell(row=row, column=5, value=round(r['min_latency'], 1))
            ws.cell(row=row, column=6, value=round(r['p99_latency'], 1))
            ws.cell(row=row, column=7, value=test_counts.get(r['acks'], 0))

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.BORDER
            row += 1

        chart_start_row = row + 2

        # =======================================================================
        # CHART 1: Throughput by acks
        # =======================================================================
        ws.cell(row=chart_start_row, column=10, value='acks')
        ws.cell(row=chart_start_row, column=11, value='Avg Throughput')

        for idx, r in acks_data.iterrows():
            ws.cell(row=chart_start_row + 1 + idx, column=10, value=str(r['acks']))
            ws.cell(row=chart_start_row + 1 + idx, column=11, value=round(r['avg_throughput'], 2))

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Throughput by acks Setting"
        chart1.style = self.CHART_STYLE
        chart1.y_axis.title = "MB/sec"
        chart1.x_axis.title = "acks value"

        data = Reference(ws, min_col=11, min_row=chart_start_row,
                        max_row=chart_start_row + len(acks_data))
        categories = Reference(ws, min_col=10, min_row=chart_start_row + 1,
                              max_row=chart_start_row + len(acks_data))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.width = 12
        chart1.height = 8

        ws.add_chart(chart1, f'A{chart_start_row}')

        # =======================================================================
        # CHART 2: Latency by acks
        # =======================================================================
        ws.cell(row=chart_start_row, column=13, value='acks')
        ws.cell(row=chart_start_row, column=14, value='Avg Latency')
        ws.cell(row=chart_start_row, column=15, value='P99 Latency')

        for idx, r in acks_data.iterrows():
            ws.cell(row=chart_start_row + 1 + idx, column=13, value=str(r['acks']))
            ws.cell(row=chart_start_row + 1 + idx, column=14, value=round(r['avg_latency'], 1))
            ws.cell(row=chart_start_row + 1 + idx, column=15, value=round(r['p99_latency'], 1))

        chart2 = BarChart()
        chart2.type = "col"
        chart2.title = "Latency by acks Setting"
        chart2.style = self.CHART_STYLE
        chart2.y_axis.title = "Latency (ms)"

        data = Reference(ws, min_col=14, max_col=15, min_row=chart_start_row,
                        max_row=chart_start_row + len(acks_data))
        categories = Reference(ws, min_col=13, min_row=chart_start_row + 1,
                              max_row=chart_start_row + len(acks_data))
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(categories)
        chart2.width = 12
        chart2.height = 8

        ws.add_chart(chart2, f'G{chart_start_row}')

        # =======================================================================
        # TRADE-OFF ANALYSIS
        # =======================================================================
        analysis_row = chart_start_row + 18
        ws[f'A{analysis_row}'] = 'Trade-off Analysis'
        ws[f'A{analysis_row}'].font = self.SECTION_FONT
        analysis_row += 1

        # Calculate performance differences
        acks0_data = acks_data[acks_data['acks'] == '0']
        acksall_data = acks_data[(acks_data['acks'] == 'all') | (acks_data['acks'] == '-1')]

        if not acks0_data.empty and not acksall_data.empty:
            acks0_throughput = acks0_data['avg_throughput'].iloc[0]
            acksall_throughput = acksall_data['avg_throughput'].iloc[0]
            acks0_latency = acks0_data['avg_latency'].iloc[0]
            acksall_latency = acksall_data['avg_latency'].iloc[0]

            throughput_diff = ((acks0_throughput - acksall_throughput) / acksall_throughput) * 100 if acksall_throughput > 0 else 0
            latency_diff = ((acksall_latency - acks0_latency) / acks0_latency) * 100 if acks0_latency > 0 else 0

            ws[f'A{analysis_row}'] = f'Throughput cost of full durability (acks=0 vs acks=all):'
            ws[f'B{analysis_row}'] = f'{throughput_diff:.1f}% faster with acks=0'
            ws[f'B{analysis_row}'].font = Font(bold=True, color='FF0000' if throughput_diff > 50 else '000000')
            analysis_row += 1

            ws[f'A{analysis_row}'] = f'Latency cost of full durability:'
            ws[f'B{analysis_row}'] = f'{latency_diff:.1f}% higher with acks=all'
            ws[f'B{analysis_row}'].font = Font(bold=True)
            analysis_row += 2

            # Recommendations
            ws[f'A{analysis_row}'] = 'RECOMMENDATIONS:'
            ws[f'A{analysis_row}'].font = Font(bold=True, size=12)
            analysis_row += 1

            if throughput_diff > 100:
                ws[f'A{analysis_row}'] = '  - Use acks=0 for non-critical telemetry/logs (>2x performance gain)'
                analysis_row += 1
                ws[f'A{analysis_row}'] = '  - Use acks=1 for balanced workloads'
                analysis_row += 1
                ws[f'A{analysis_row}'] = '  - Use acks=all only for mission-critical financial/transactional data'
            elif throughput_diff > 50:
                ws[f'A{analysis_row}'] = '  - Consider acks=1 for balanced performance and durability'
                analysis_row += 1
                ws[f'A{analysis_row}'] = '  - acks=all recommended for data where loss is unacceptable'
            else:
                ws[f'A{analysis_row}'] = '  - Performance difference is minimal - use acks=all for safety'
                analysis_row += 1
                ws[f'A{analysis_row}'] = '  - Low overhead for full durability in your environment'

        self._auto_width_columns(ws)

    def create_raw_data_sheet(self) -> None:
        """Create Raw Data sheet with all results for pivot tables."""
        ws = self.wb.create_sheet('Raw Data')
        self.log("Creating Raw Data sheet...")

        # Combine all results into flat structure
        all_rows = []

        for r in self.producer_results:
            config = r.get('configuration', {})
            metrics = r.get('metrics', {})
            all_rows.append({
                'test_type': 'producer',
                'scenario': r.get('scenario', ''),
                'acks': config.get('acks', ''),
                'batch_size': config.get('batch_size', ''),
                'linger_ms': config.get('linger_ms', ''),
                'compression_type': config.get('compression_type', config.get('compression', '')),
                'record_size': config.get('record_size', ''),
                'num_producers': config.get('num_producers', 1),
                'records_sent': metrics.get('records_sent', ''),
                'throughput_rps': metrics.get('throughput_rps', ''),
                'throughput_mb': metrics.get('throughput_mb', ''),
                'avg_latency_ms': metrics.get('avg_latency_ms', ''),
                'max_latency_ms': metrics.get('max_latency_ms', ''),
                'p50_ms': metrics.get('p50_ms', ''),
                'p95_ms': metrics.get('p95_ms', ''),
                'p99_ms': metrics.get('p99_ms', ''),
                'p999_ms': metrics.get('p999_ms', ''),
            })

        for r in self.consumer_results:
            config = r.get('configuration', {})
            metrics = r.get('metrics', {})
            all_rows.append({
                'test_type': 'consumer',
                'scenario': r.get('scenario', ''),
                'fetch_min_bytes': config.get('fetch_min_bytes', ''),
                'max_poll_records': config.get('max_poll_records', ''),
                'num_consumers': config.get('num_consumers', 1),
                'throughput_mb': metrics.get('throughput_mb_sec', ''),
                'throughput_rps': metrics.get('throughput_msg_sec', ''),
                'data_consumed_mb': metrics.get('data_consumed_mb', ''),
                'rebalance_time_ms': metrics.get('rebalance_time_ms', ''),
                'fetch_time_ms': metrics.get('fetch_time_ms', ''),
            })

        if not all_rows:
            ws['A1'] = 'No test data available'
            return

        df = pd.DataFrame(all_rows)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                cell.border = self.BORDER

        self._auto_width_columns(ws)

    # ==========================================================================
    # SHEET 10: RECOMMENDATIONS
    # ==========================================================================

    def create_recommendations_sheet(self) -> None:
        """
        Create Recommendations sheet with scored configurations by use case.

        Includes (per PRD):
        1. Maximum Throughput configuration
        2. Balanced Performance configuration
        3. Mission-Critical (Durability) configuration
        4. Scoring methodology explanation
        5. When to adjust settings guide
        """
        ws = self.wb.create_sheet('Recommendations')
        self.log("Creating Recommendations sheet...")

        if self.producer_df.empty:
            ws['A1'] = 'No test results available for recommendations'
            return

        # Title
        ws['A1'] = 'RECOMMENDED CONFIGURATIONS BY USE CASE'
        ws['A1'].font = self.TITLE_FONT
        ws.merge_cells('A1:H1')

        ws['A2'] = 'Optimal configurations identified from your test results'
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:H2')

        scored_df = self.scores.get('scored_df', self.producer_df)
        row = 4

        # =======================================================================
        # USE CASE 1: MAXIMUM THROUGHPUT
        # =======================================================================
        ws[f'A{row}'] = 'USE CASE 1: MAXIMUM THROUGHPUT'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        if self.scores.get('best_throughput_idx') is not None:
            best = scored_df.loc[self.scores['best_throughput_idx']]
            self._write_recommendation_box(ws, row, best, 'throughput')
            row += 12

        # =======================================================================
        # USE CASE 2: BALANCED PERFORMANCE
        # =======================================================================
        ws[f'A{row}'] = 'USE CASE 2: BALANCED PERFORMANCE'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        if self.scores.get('best_balanced_idx') is not None:
            best = scored_df.loc[self.scores['best_balanced_idx']]
            self._write_recommendation_box(ws, row, best, 'balanced')
            row += 12

        # =======================================================================
        # USE CASE 3: MISSION-CRITICAL (DURABILITY)
        # =======================================================================
        ws[f'A{row}'] = 'USE CASE 3: MISSION-CRITICAL (DURABILITY FIRST)'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        # Find best config with acks=all or acks=-1
        durable_configs = scored_df[(scored_df['acks'] == 'all') | (scored_df['acks'] == '-1')]
        if not durable_configs.empty:
            best_durable_idx = durable_configs['score_durability'].idxmax()
            best = scored_df.loc[best_durable_idx]
            self._write_recommendation_box(ws, row, best, 'durability')
            row += 12
        else:
            ws[f'A{row}'] = 'No tests with acks=all found. Run tests with acks=all for durability recommendations.'
            row += 2

        # =======================================================================
        # SCORING METHODOLOGY
        # =======================================================================
        ws[f'A{row}'] = 'SCORING METHODOLOGY'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 2

        formulas = [
            'Throughput Score = (test_throughput / max_observed) * 100',
            'Latency Score = 100 - ((test_latency / max_observed) * 100)',
            'Consistency Score = Based on P99/Avg latency ratio (lower variance = higher score)',
            '',
            'Weighted Scores by Use Case:',
            '  Max Throughput:  40% throughput + 20% latency + 40% consistency',
            '  Balanced:        30% throughput + 35% latency + 35% consistency',
            '  Durability:      20% throughput + 35% latency + 45% consistency',
        ]

        for formula in formulas:
            ws[f'A{row}'] = formula
            row += 1

        row += 1

        # =======================================================================
        # WHEN TO ADJUST SETTINGS
        # =======================================================================
        ws[f'A{row}'] = 'WHEN TO ADJUST THESE SETTINGS'
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')
        row += 2

        adjustments = [
            ('IF YOU OBSERVE:', 'ADJUST:'),
            ('Low throughput, high P99 latency', 'Increase batch.size'),
            ('Good throughput but spiky latency', 'Increase linger.ms'),
            ('High CPU on producer', 'Reduce compression or increase batch.size'),
            ('Message loss in production', 'Increase acks level'),
            ('Replication lag too high', 'Reduce num.producers or increase broker replicas'),
        ]

        for observe, adjust in adjustments:
            ws[f'A{row}'] = observe
            ws[f'A{row}'].font = Font(bold=True) if observe == 'IF YOU OBSERVE:' else None
            ws[f'D{row}'] = adjust
            ws[f'D{row}'].font = Font(bold=True) if adjust == 'ADJUST:' else None
            row += 1

        self._auto_width_columns(ws)

    def _write_recommendation_box(self, ws, start_row: int, config, use_case: str) -> None:
        """Write a recommendation box for a specific use case."""
        row = start_row

        # Configuration
        ws[f'A{row}'] = 'Configuration:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = f"  acks = {config['acks']}"
        row += 1
        ws[f'A{row}'] = f"  batch.size = {config['batch_size']}"
        row += 1
        ws[f'A{row}'] = f"  linger.ms = {config['linger_ms']}"
        row += 1
        ws[f'A{row}'] = f"  compression.type = {config['compression_type']}"
        row += 1
        ws[f'A{row}'] = f"  record.size = {config['record_size']}"
        row += 1

        # Expected Performance
        row += 1
        ws[f'A{row}'] = 'Expected Performance:'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = f"  Throughput: {config['throughput_mb']:.2f} MB/sec ({config['throughput_rps']:,.0f} records/sec)"
        ws[f'A{row}'].fill = self.GOOD_FILL if config['throughput_mb'] > 50 else None
        row += 1
        ws[f'A{row}'] = f"  Avg Latency: {config['avg_latency_ms']:.0f} ms"
        row += 1
        ws[f'A{row}'] = f"  P99 Latency: {config['p99_ms']:.0f} ms"
        row += 1

        # Score
        score_col = f'score_{use_case}' if use_case != 'throughput' else 'score_max_throughput'
        if score_col in config:
            ws[f'A{row}'] = f"  Score: {config[score_col]:.0f}/100"
            ws[f'A{row}'].font = Font(bold=True)


def main():
    """Main entry point for the report generator."""
    parser = argparse.ArgumentParser(
        description='Generate comprehensive Excel reports from Kafka performance test results (PRD v2)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s parsed_results.json kafka_report.xlsx
  %(prog)s ./results/parsed_data/results.json ./results/reports/report.xlsx
  %(prog)s -i parsed.json -o report.xlsx --verbose

Output:
  Creates a 10-sheet Excel workbook with 30+ visualizations including:
  - Dashboard with KPI summary and key charts
  - THE CRITICAL "knee" chart showing system saturation
  - Configuration heatmaps for parameter optimization
  - Scored recommendations by use case
        """
    )

    parser.add_argument(
        'input_json',
        nargs='?',
        default='./results/parsed_data/parsed_results.json',
        help='Parsed JSON results file (default: ./results/parsed_data/parsed_results.json)'
    )

    parser.add_argument(
        'output_xlsx',
        nargs='?',
        default='./results/reports/kafka_perf_report.xlsx',
        help='Output Excel file path (default: ./results/reports/kafka_perf_report.xlsx)'
    )

    parser.add_argument(
        '--input', '-i',
        dest='input_opt',
        help='Alternative way to specify input JSON file'
    )

    parser.add_argument(
        '--output', '-o',
        dest='output_opt',
        help='Alternative way to specify output Excel file'
    )

    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output'
    )

    args = parser.parse_args()

    # Handle alternative argument forms
    input_json = args.input_opt or args.input_json
    output_xlsx = args.output_opt or args.output_xlsx

    # Generate report
    reporter = KafkaPerformanceReporter(input_json, output_xlsx, verbose=args.verbose)
    output_file = reporter.generate_report()

    print(f"\n{'='*60}")
    print(f"Kafka Performance Report Generated Successfully!")
    print(f"{'='*60}")
    print(f"Output: {output_file}")
    print(f"\n10 Sheets Created (PRD v2 Specification):")
    print("  1. Dashboard          - Executive summary with 4 key charts")
    print("  2. Throughput Analysis - Deep dive with 4 charts")
    print("  3. Latency Analysis   - Percentile distributions with 4 charts")
    print("  4. Trade-off Analysis - THE KNEE CHART + score breakdown")
    print("  5. Scaling Performance - Degradation analysis with 4 charts")
    print("  6. Configuration Heatmap - Multi-dimensional parameter analysis")
    print("  7. Message Size Impact - Size efficiency curves")
    print("  8. Acks Comparison    - Durability vs performance trade-offs")
    print("  9. Raw Data           - Complete data for pivot tables")
    print(" 10. Recommendations    - Scored configs by use case")
    print(f"\nTotal visualizations: 30+ charts and heatmaps")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
